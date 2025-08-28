# 松戸市の全死因死者数を2021年下半期～2024年上半期までワクチン接種回数別に出力するプログラム
from openpyxl import load_workbook
import csv
from datetime import datetime
import numpy as np

# 入出力ファイル
input_file = "matsudo.xlsx"  # 読み込み元 Excel（松戸のデータ）
#OUT = "death_mat65-89.csv" #65～89歳
#OUT = "death_mat50-64.csv" #50～64歳
OUT = "death_mat20-49.csv" #20～49歳

# カウンタ初期化
data = [0] * 8  # data[0]～data[7]

date = [0] * 7
death_count = np.zeros((8, 6), dtype=int)

date[0] = datetime.strptime("2021/7/1", "%Y/%m/%d")
date[1] = datetime.strptime("2022/1/1", "%Y/%m/%d")
date[2] = datetime.strptime("2022/7/1", "%Y/%m/%d")
date[3] = datetime.strptime("2023/1/1", "%Y/%m/%d")
date[4] = datetime.strptime("2023/7/1", "%Y/%m/%d")
date[5] = datetime.strptime("2024/1/1", "%Y/%m/%d")
date[6] = datetime.strptime("2024/7/1", "%Y/%m/%d")

def is_number(val):
    if val is None or str(val).strip() == "":
        return False
    return str(val).isdigit()

def is_empty(val):
    return val is None or str(val).strip() == ""
    
# Excel 読み込み
wb = load_workbook(input_file, data_only=True)
ws = wb.active  # 最初のシートを使用

# 全行を2次元リスト化（インデックス合わせのため 0-based にする）
rows = list(ws.iter_rows(values_only=True))

# 全行を走査
for i, row in enumerate(rows):
    # None 回避のため空文字変換
    row = [str(cell) if cell is not None else "" for cell in row]

#    if row[0] not in ("1956～1960", "1951～1955", "1946～1950", "1941～1945", "1936～1940"): #65～89歳
#    if row[0] not in ("1971～1975", "1966～1970", "1961～1965"): #50～64歳
    if row[0] not in ("2001～2005", "1996～2000", "1991～1995", "1986～1990",  "1981～1985", "1976～1980"): #20～49歳

        continue

    if is_number(row[3]): # 死亡日記載あり
        if is_empty(row[4]): # 0回接種者
            death_date = datetime.strptime(row[3], "%Y%m%d")
            for j in range(6):
                if (((death_date - date[j]).days >= 0) & ((date[j+1] - death_date).days > 0)):
                        death_count[0][j] += 1

        for n in range(0, 7):  # 1-7回接種者
            if is_number(row[4 + 3*n]) and is_empty(row[4 + 3*n + 3]):
                death_date = datetime.strptime(row[3], "%Y%m%d")
                for j in range(6):
                    if ((death_date - date[j]).days >= 0) & ((date[j+1] - death_date).days > 0):
                        death_count[n+1][j] += 1

with open(OUT, "w") as o:
    for i in range(8):
        for j in range(6):
            print(death_count[i][j], "," , end="", file=o)
        print("", file=o)